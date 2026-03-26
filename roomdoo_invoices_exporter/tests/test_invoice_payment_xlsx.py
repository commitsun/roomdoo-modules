# Copyright 2026 Roomdoo - Commit[Sun]
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl).

import unittest
from io import BytesIO

from odoo import Command
from odoo.exceptions import UserError
from odoo.tests import tagged

from odoo.addons.account.tests.common import AccountTestInvoicingCommon

try:
    from openpyxl import load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@unittest.skipUnless(HAS_OPENPYXL, "openpyxl not available")
@tagged("post_install", "-at_install")
class TestInvoicePaymentXlsx(AccountTestInvoicingCommon):
    @classmethod
    def setUpClass(cls, chart_template_ref=None):
        super().setUpClass(chart_template_ref=chart_template_ref)

        cls.partner_a.write(
            {
                "vat": "ES12345678A",
                "street": "Calle Test 1",
                "city": "Madrid",
                "zip": "28001",
                "country_id": cls.env.ref("base.es").id,
            }
        )

        cls.bank_journal = cls.company_data["default_journal_bank"]

        # Invoice with two payments (partially paid: 400 + 300 of 1000)
        cls.invoice_multi_pay = cls.env["account.move"].create(
            {
                "move_type": "out_invoice",
                "partner_id": cls.partner_a.id,
                "invoice_date": "2025-01-15",
                "invoice_line_ids": [
                    Command.create(
                        {
                            "name": "Product A",
                            "quantity": 1,
                            "price_unit": 1000.0,
                        }
                    ),
                ],
            }
        )
        cls.invoice_multi_pay.action_post()

        for amount, pay_date in [(400.0, "2025-01-20"), (300.0, "2025-01-25")]:
            wizard = (
                cls.env["account.payment.register"]
                .with_context(
                    active_model="account.move",
                    active_ids=cls.invoice_multi_pay.ids,
                )
                .create(
                    {
                        "amount": amount,
                        "journal_id": cls.bank_journal.id,
                        "payment_date": pay_date,
                    }
                )
            )
            wizard._create_payments()

        # Invoice without payments
        cls.invoice_no_pay = cls.env["account.move"].create(
            {
                "move_type": "out_invoice",
                "partner_id": cls.partner_a.id,
                "invoice_date": "2025-02-01",
                "invoice_line_ids": [
                    Command.create(
                        {
                            "name": "Product B",
                            "quantity": 1,
                            "price_unit": 500.0,
                        }
                    ),
                ],
            }
        )
        cls.invoice_no_pay.action_post()

        # Overdue invoice (date in the past, no payments)
        cls.invoice_overdue = cls.env["account.move"].create(
            {
                "move_type": "out_invoice",
                "partner_id": cls.partner_a.id,
                "invoice_date": "2024-01-01",
                "invoice_line_ids": [
                    Command.create(
                        {
                            "name": "Product C",
                            "quantity": 1,
                            "price_unit": 200.0,
                        }
                    ),
                ],
            }
        )
        cls.invoice_overdue.action_post()

    # -----------------------------------------------------------------
    # Helpers
    # -----------------------------------------------------------------

    def _generate_report(self, invoices):
        report_model = self.env[
            "report.roomdoo_invoices_exporter.invoice_payment_report"
        ].with_context(active_model="account.move")
        return report_model.create_xlsx_report(invoices.ids, data=None)

    def _load_workbook(self, content):
        return load_workbook(BytesIO(content))

    def _find_summary_value(self, ws, label_substring):
        for row in ws.iter_rows(min_row=1, max_col=2):
            cell_label = row[0].value
            if cell_label and label_substring in str(cell_label):
                return row[1].value
        return None

    # -----------------------------------------------------------------
    # Tests
    # -----------------------------------------------------------------

    def test_report_generates_three_sheets(self):
        content, fmt = self._generate_report(
            self.invoice_multi_pay | self.invoice_no_pay
        )
        self.assertEqual(fmt, "xlsx")
        wb = self._load_workbook(content)
        self.assertEqual(len(wb.worksheets), 3)

    def test_invoice_with_multiple_payments(self):
        """One invoice row, N payment rows."""
        content, _ = self._generate_report(self.invoice_multi_pay)
        wb = self._load_workbook(content)
        ws_inv = wb.worksheets[0]
        ws_pay = wb.worksheets[1]
        # 1 header + 1 data row
        self.assertEqual(ws_inv.max_row, 2)
        # 1 header + 2 payment rows
        self.assertEqual(ws_pay.max_row, 3)

    def test_invoice_without_payments(self):
        """Invoice present in sheet 1, absent from sheet 2."""
        content, _ = self._generate_report(self.invoice_no_pay)
        wb = self._load_workbook(content)
        ws_inv = wb.worksheets[0]
        ws_pay = wb.worksheets[1]
        self.assertEqual(ws_inv.max_row, 2)
        self.assertEqual(ws_pay.max_row, 1)

    def test_summary_totals(self):
        """Verify key aggregates in the summary sheet."""
        invoices = self.invoice_multi_pay | self.invoice_no_pay
        content, _ = self._generate_report(invoices)
        wb = self._load_workbook(content)
        ws_summary = wb.worksheets[2]

        total_invoiced = self._find_summary_value(ws_summary, "Total invoiced")
        total_collected = self._find_summary_value(ws_summary, "Total collected")
        total_pending = self._find_summary_value(ws_summary, "Total pending")

        expected_invoiced = 1000.0 + 500.0
        expected_collected = 400.0 + 300.0
        expected_pending = expected_invoiced - expected_collected

        self.assertAlmostEqual(total_invoiced, expected_invoiced, places=2)
        self.assertAlmostEqual(total_collected, expected_collected, places=2)
        self.assertAlmostEqual(total_pending, expected_pending, places=2)

        diff = self._find_summary_value(ws_summary, "Difference")
        self.assertAlmostEqual(diff, 0.0, places=2)

    def test_overdue_invoice(self):
        """Overdue invoice shows correct status and amount."""
        content, _ = self._generate_report(self.invoice_overdue)
        wb = self._load_workbook(content)
        ws_inv = wb.worksheets[0]
        # Payment Status column = 18 (1-based in openpyxl)
        status = ws_inv.cell(row=2, column=18).value
        self.assertEqual(status, "Overdue")
        # Overdue Amount column = 17
        overdue_amount = ws_inv.cell(row=2, column=17).value
        self.assertAlmostEqual(overdue_amount, 200.0, places=2)

    def test_multi_company_raises_error(self):
        """Exporting invoices from different companies must fail."""
        invoice_company_2 = (
            self.env["account.move"]
            .with_company(self.company_data_2["company"])
            .create(
                {
                    "move_type": "out_invoice",
                    "partner_id": self.partner_a.id,
                    "invoice_date": "2025-03-01",
                    "invoice_line_ids": [
                        Command.create(
                            {
                                "name": "Other",
                                "quantity": 1,
                                "price_unit": 100.0,
                            }
                        ),
                    ],
                }
            )
        )
        invoice_company_2.action_post()
        with self.assertRaises(UserError):
            self._generate_report(self.invoice_multi_pay | invoice_company_2)
